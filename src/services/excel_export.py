import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from openpyxl.utils import get_column_letter

def generate_olympiad_report(olympiad, applications):
    """Генерирует Excel-отчет по заявкам на олимпиаду"""
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    # Заголовок отчета
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Отчет по заявкам на олимпиаду: {olympiad['title']}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Информация об олимпиаде
    ws.append(["Организатор:", olympiad['organizer']])
    ws.append(["Дисциплина:", olympiad['subject_title']])
    ws.append(["Дата начала:", olympiad['start_date'].strftime("%d.%m.%Y")])
    ws.append(["Дата окончания:", olympiad['end_date'].strftime("%d.%m.%Y")])
    ws.append([])
    
    # Заголовки таблицы заявок
    headers = [
        "ID заявки", 
        "Фамилия", 
        "Имя", 
        "Отчество", 
        "Статус", 
        "Дата подачи"
    ]
    ws.append(headers)
    
    # Стили для заголовков
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    border = Border(bottom=Side(style='medium'))
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=7, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Заполняем данными
    for app in applications:
        ws.append([
            app['application_id'],
            app['last_name'],
            app['first_name'],
            app.get('middle_name', ''),
            app['status_name'],
            app['created_date'].strftime("%d.%m.%Y %H:%M")
        ])
    
    # Добавляем итоговую строку
    last_row = ws.max_row + 1
    ws.merge_cells(f'A{last_row}:D{last_row}')
    total_cell = ws.cell(row=last_row, column=1)
    total_cell.value = f"Всего заявок: {len(applications)}"
    total_cell.font = Font(bold=True)
    
    # Автонастройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_olympiad_{olympiad['olympiad_id']}_{timestamp}.xlsx"
    wb.save(filename)
    
    return filename